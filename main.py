import dearpygui.dearpygui as dpg
from sqlalchemy import String, Column,Text, Integer,MetaData,engine, text
import sqlalchemy as sql
import re
import psycopg2 
import dearpygui.demo as demo
import openpyxl as op

engine = sql.create_engine('postgresql://postgres:1111@localhost:5432/postgres',echo=True)
connection = engine.connect()
metadata = MetaData()
metadata.create_all(engine)

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION insert_into_gr_proj(
    in_codkon TEXT,
    in_g1 INTEGER,
    in_g8 TEXT DEFAULT NULL,
    in_g7 TEXT DEFAULT '00.00',
    in_z2 TEXT DEFAULT NULL,
    in_g5 INTEGER DEFAULT 0,
    in_codvuz INTEGER DEFAULT NULL,
    in_g6 TEXT DEFAULT NULL,
    in_g9 TEXT DEFAULT NULL,
    in_g10 TEXT DEFAULT NULL,
    in_g11 TEXT DEFAULT NULL)
RETURNS VOID AS $$
BEGIN
    INSERT INTO gr_proj (codkon,g1,g8,g7,z2,g5,g2,g21,g22,g23,g24,codvuz,g6,g9,g10,g11) VALUES (in_codkon,in_g1,in_g8,in_g7,in_z2,in_g5,0,0,0,0,0,in_codvuz,in_g6,in_g9,in_g10,in_g11);
END;
$$ LANGUAGE plpgsql;
"""
))

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION edit_row_gr_proj(
    in_codkon TEXT,
    in_g1 INTEGER,
    in_g8 TEXT DEFAULT NULL,
    in_g7 TEXT DEFAULT NULL,
    in_z2 TEXT DEFAULT NULL,
    in_g5 INTEGER DEFAULT NULL,
    in_codvuz INTEGER DEFAULT NULL,
    in_g6 TEXT DEFAULT NULL,
    in_g9 TEXT DEFAULT NULL,
    in_g10 TEXT DEFAULT NULL,
    in_g11 TEXT DEFAULT NULL,
    out_codkon TEXT DEFAULT NULL,
    out_g1 INTEGER DEFAULT NULL,
    out_g8 TEXT DEFAULT NULL,
    out_g7 TEXT DEFAULT NULL,
    out_z2 TEXT DEFAULT NULL,
    out_g5 INTEGER DEFAULT NULL,
    out_codvuz INTEGER DEFAULT NULL,
    out_g6 TEXT DEFAULT NULL,
    out_g9 TEXT DEFAULT NULL,
    out_g10 TEXT DEFAULT NULL,
    out_g11 TEXT DEFAULT NULL
    )
RETURNS VOID AS $$
BEGIN
    UPDATE gr_proj SET codkon = COALESCE(out_codkon, in_codkon), g1 = COALESCE(out_g1,in_g1), g8 = COALESCE(out_g8, in_g8),
    g7 = COALESCE(out_g7, in_g7), z2 = COALESCE(out_z2, in_z2),g5 = COALESCE(out_g5, in_g5), codvuz = COALESCE(out_codvuz, in_codvuz),
    g6 = COALESCE(out_g6, in_g6), g9 = COALESCE(out_g9, in_g9), g10 = COALESCE(out_g10, in_g10),g11 = COALESCE(out_g11, in_g11) 
    WHERE codkon = in_codkon AND g1 = in_g1;
END;
$$ LANGUAGE plpgsql;
"""
))

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION delete_row_gr_proj(
    in_codkon TEXT,
    in_g1 INTEGER
)
RETURNS VOID AS $$
BEGIN
    DELETE FROM gr_proj WHERE codkon=in_codkon AND g1 = in_g1;
END;
$$ LANGUAGE plpgsql;
"""
))

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION planed_financing()
RETURNS TRIGGER AS $$
DECLARE 
array_str text [] := ARRAY['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17'];
i text;
BEGIN
    FOREACH i IN ARRAY array_str
        LOOP
        UPDATE gr_konk SET k12 = COALESCE((SELECT SUM(COALESCE(g5, 0)) FROM gr_proj WHERE codkon=i),0) WHERE codkon=i;
        END LOOP;
        RETURN NEW; 
END;
$$ LANGUAGE plpgsql;
"""
))

connection.execute(text(
"""
CREATE OR REPLACE TRIGGER gr_konk_financing
AFTER INSERT OR DELETE OR UPDATE OF g5
ON gr_proj
FOR EACH ROW
EXECUTE PROCEDURE planed_financing();
"""
))

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION konk_participants()
RETURNS TRIGGER AS $$
DECLARE 
array_str text [] := ARRAY['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17'];
i text;
BEGIN
    FOREACH i IN ARRAY array_str
        LOOP
        UPDATE gr_konk SET npr = COALESCE((SELECT COALESCE(COUNT(*),0) FROM gr_proj WHERE codkon=i),0) WHERE codkon=i;
        END LOOP;
        RETURN NEW; 
END;
$$ LANGUAGE plpgsql;
"""
))

connection.execute(text(
"""
CREATE OR REPLACE TRIGGER gr_konk_npr
AFTER INSERT OR DELETE  
ON gr_proj
FOR EACH ROW
EXECUTE PROCEDURE konk_participants();
"""
))

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION gr_proj_kv_financing(
    column_name TEXT,
    portion INTEGER
)
RETURNS VOID AS $$
BEGIN
    EXECUTE format('UPDATE gr_proj SET %I = %I + %L', column_name,column_name,portion);
END;
$$ LANGUAGE plpgsql;
"""
))

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION gr_konk_actual_financing()
RETURNS void AS $$
BEGIN
    UPDATE gr_konk SET k4 = k41 + k42 + k43 + k44;
END;
$$ LANGUAGE plpgsql;
"""
))

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION kv1()
RETURNS void AS $$
DECLARE 
array_str text [] := ARRAY['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17'];
i text;
BEGIN
    FOREACH i IN ARRAY array_str
        LOOP
            UPDATE gr_konk SET k41 = COALESCE((SELECT SUM(COALESCE(g21, 0)) FROM gr_proj WHERE codkon=i),0) WHERE codkon=i;
        END LOOP;
END;
$$ LANGUAGE plpgsql;
"""
))

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION kv2()
RETURNS void AS $$
DECLARE 
array_str text [] := ARRAY['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17'];
i text;
BEGIN
    FOREACH i IN ARRAY array_str
        LOOP
            UPDATE gr_konk SET k42 = COALESCE((SELECT SUM(COALESCE(g22, 0)) FROM gr_proj WHERE codkon=i),0) WHERE codkon=i;
        END LOOP;
END;
$$ LANGUAGE plpgsql;
"""
))

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION kv3()
RETURNS void AS $$
DECLARE 
array_str text [] := ARRAY['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17'];
i text;
BEGIN
    FOREACH i IN ARRAY array_str
        LOOP
        UPDATE gr_konk SET k43 = COALESCE((SELECT SUM(COALESCE(g23, 0)) FROM gr_proj WHERE codkon=i),0) WHERE codkon=i;
        END LOOP;
END;
$$ LANGUAGE plpgsql;
"""
))

connection.execute(text(
"""
CREATE OR REPLACE FUNCTION kv4()
RETURNS void AS $$
DECLARE 
array_str text [] := ARRAY['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17'];
i text;
BEGIN
    FOREACH i IN ARRAY array_str
        LOOP
        UPDATE gr_konk SET k44 = COALESCE((SELECT SUM(COALESCE(g24, 0)) FROM gr_proj WHERE codkon=i),0) WHERE codkon=i;
        END LOOP;
END;
$$ LANGUAGE plpgsql;
"""
))


connection.execute(text(
"""
CREATE OR REPLACE FUNCTION gr_proj_actual_fin()
RETURNS VOID AS $$
BEGIN
    UPDATE gr_proj SET g2 = g21 + g22 + g23 + g24;
END;
$$ LANGUAGE plpgsql;
"""
))


dpg.create_context()
dpg.create_viewport(title= 'Support of competitions for grants', width = 2200, height = 1200)


big_let_start = 0x00C0  # Capital "A" in cyrillic alphabet
big_let_end = 0x00DF  # Capital "Я" in cyrillic alphabet
small_let_end = 0x00FF  # small "я" in cyrillic alphabet
remap_big_let = 0x0410  # Starting number for remapped cyrillic alphabet
alph_len = big_let_end - big_let_start + 1  # adds the shift from big letters to small
alph_shift = remap_big_let - big_let_start  # adds the shift from remapped to non-remapped

#demo.show_demo()

def to_cyr(instr):  # conversion function
        out = []  # start with empty output
        for i in range(0, len(instr)):  # cycle through letters in input string
            if ord(instr[i]) in range(big_let_start, small_let_end + 1):  # check if the letter is cyrillic
                out.append(chr(ord(instr[i]) + alph_shift))  # if it is change it and add to output list
            else:
                out.append(instr[i])  # if it isn`t don`t change anything and just add it to output list
        return ''.join(out)  # convert list to string

with dpg.font_registry():
    with dpg.font("NotoMono-Regular.ttf", 20) as font1:
        dpg.add_font_range_hint(dpg.mvFontRangeHint_Cyrillic)
        dpg.add_font_chars([0x2116])
        dpg.bind_font(font1)       
        biglet = remap_big_let  # Starting number for remapped cyrillic alphabet
        for i1 in range(big_let_start, big_let_end + 1):  # Cycle through big letters in cyrillic alphabet
            dpg.add_char_remap(i1, biglet)  # Remap the big cyrillic letter
            dpg.add_char_remap(i1 + alph_len, biglet + alph_len)  # Remap the small cyrillic letter
            biglet += 1  # choose next letter

table_list_tag = ['GR_KONK','GR_PROJ','VUZ']
table_list_name_size = [['gr_konk',9],['gr_proj',16],['vuz',11]]
GR_KONK_label_list = [' ', 'k2', 'codkon', 'npr', 'k12', 'k4', 'k41', 'k42', 'k43', 'k44']
GR_PROJ_label_list = [' ', 'codkon', 'g1','g8','g7','z2','g5','g2','g21','g22','g23','g24','codvuz','g6','g9','g10','g11']
VUZ_label_list = [' ', 'codvuz', 'z1', 'z1full','z2','region', 'city', 'status', 'obl','oblname','gr_ved','prof']
window_specs = [['GR_KONK',0], ['GR_PROJ',1], ['VUZ',2]]
GR_KONK_label_list_ru = [' ', 'Конкурс', 'Код конкурса', 'кол-во НИР по грантам', 'Плановый объем гранта', 'Фактический объем гранта', '1-ый кв', '2-ой кв', '3-ий кв', '4-ый кв']
GR_PROJ_label_list_ru  = [' ','Код конкурса', 'Код НИР', 'Руководитель НИР', 'ГРНТИ','ВУЗ', 'Плановый объем гранта', 'Фактический объем гранта',
 '1-ый кв', '2-ой кв', '3-ий кв', '4-ый кв', 'Код вуза', 'НИР',  'Должность', 'Ученое звание', 'Ученая степень']
VUZ_label_list_ru  = [' ', 'Код вуза', 'ВУЗ', 'Юр наз-ие ВУЗа','Наименование','Фед. округ', 'Город', 'Статус', 'Субъект','Область','gr_ved','Направленность']


VUZ_analys_sql_string = """SELECT ROW_NUMBER() OVER (ORDER BY gp.z2 ASC), gp.z2, COUNT(g1), SUM(g5), COUNT(DISTINCT gp.codkon) 
                                            FROM gr_proj gp
                                            INNER JOIN gr_konk gk  ON gk.codkon = gp.codkon
                                            INNER JOIN vuz vz ON vz.codvuz=gp.codvuz
                                            GROUP BY gp.z2"""

KONK_analys_sql_string = """SELECT ROW_NUMBER() OVER (ORDER BY npr DESC), k2, COUNT(gp.g1), SUM(gp.g5), COUNT(DISTINCT gp.z2)   
                                            FROM gr_konk gk 
                                            INNER JOIN gr_proj gp ON gk.codkon = gp.codkon
                                            INNER JOIN vuz vz ON vz.codvuz=gp.codvuz
                                            GROUP BY gp.codkon, k2, npr, k12"""

FED_analys_sql_string = """SELECT ROW_NUMBER() OVER (ORDER BY oblname ASC), 
                                            oblname,
                                            COUNT(gp.g1),
                                            (SUM(gp.g5)),
                                            COUNT(DISTINCT vz.codvuz),
                                            COUNT(DISTINCT gp.codkon)
                                            FROM vuz vz 
                                            INNER JOIN gr_proj gp ON vz.codvuz=gp.codvuz
                                            INNER JOIN gr_konk gk ON gk.codkon = gp.codkon
                                            GROUP BY oblname"""

TOTAL_for_vuz_konk_analys_sql_string = """SELECT ' ','ИТОГО', COUNT(gp.g1), SUM(gp.g5), ' '
                                            FROM gr_konk gk 
                                            INNER JOIN gr_proj gp ON gk.codkon = gp.codkon
                                            INNER JOIN vuz vz ON vz.codvuz=gp.codvuz"""

TOTAL_for_fed_analys_sql_string = '''SELECT ' ', 'ИТОГ', COUNT(gp.g1), SUM(gp.g5), (
                                    SELECT SUM(unique_values) FROM (
                                        SELECT COUNT(DISTINCT vz.codvuz) AS unique_values 
                                        FROM vuz vz
                                        INNER JOIN gr_proj gp ON vz.codvuz = gp.codvuz
                                        INNER JOIN gr_konk gk ON gp.codkon = gk.codkon
                                        GROUP BY oblname
                                    ) AS subquery
                                    ) AS sum_of_unique_values
                                    FROM gr_konk gk 
                                    INNER JOIN gr_proj gp ON gk.codkon = gp.codkon
                                    INNER JOIN vuz vz ON vz.codvuz = gp.codvuz'''


VUZ_z2_codvuz_dict = {}

result = connection.execute(text("SELECT DISTINCT z2,codvuz FROM vuz"))
for pairs in result:
    VUZ_z2_codvuz_dict[pairs[0]] = pairs[1]

table_name_id = []
high_list = []
high_list_copy = []
res_mas = []
specs_filter_list = []
filter_mas_1 = []
combo_id_mas = []
VUZ_row_num_list = []
GR_KONK_row_num_list = []
GR_PROJ_row_num_list = []
filter_string = ''
new_formed_filter_string = ''
codkon_list = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17']
codkon_names_list = ['машиностроения(01)','охраны окружающей среды и экологии человека(02)','лесного комплекса(03)','энергетики и электротехники(04)','геодезии и картографии(05)',
'информатики, кибернетики, автоматики и телемеханики,метрологии и связи(06)','фундаментального естествознания(07)', 'технологических проблем производства авиакосмическойтехникии(08)',
'сельхозмашиностроения(09)','математики(10)', 'электроники и радиотехники(11)', 'горных наук(12)', 'геологии(13)', 'приборостроения(14)', 'металлургии(15)', 'гуманитарных наук(16)',
'ядерной техники и физики пучков ионизирующих излучений(17)']
gr_proj_for_vuz_dict = {'g5':'k12','g2':'k4','g21':'k41','g22':'k42','g23':'k43','g24':'k44'}
edit_id_input = []
add_row_input = []
add_row_id = []
add_grnti = []
edit_grnti = []
first_grnti_input = 0
first_grnti_edit = 0
add_row_max_g1 = 0
grnti_edit_defaults = []
filter_condition = False
vuz_ex_name = '0'
konk_ex_name = '0'
fed_ex_name = '0'
finance_ex_name = '0'
finance_items = ['1 квартал','2 квартал','3 квартал','4 квартал']
finance_id = []
planned_financing = 1
actual_financing = 0
current_int = 0
current_procent = 0



def court_to_string(mas, result):
    for row in result:
        mas.append([str(x) for x in row])
    for c in range(len(mas)):
        mas[c] = mas[c][0]

def add_row_window_input_creation():
    for i in range(17):
        if i in [0,7,8,9,10,11,12]:
            continue
        else:
            if i == 1:
                add_row_id_1 = dpg.add_combo(parent ='ADD_ROW',width = 400, items = codkon_names_list, label = 'Конкурс', callback = add_row_text_input)
            elif i == 2:
                add_row_id_1 = dpg.add_input_text(parent ='ADD_ROW',width = 400, label = GR_PROJ_label_list_ru[i], callback = add_row_text_input, decimal=True, on_enter=True , enabled  = False)
            elif i == 4:
                add_row_id_1 = dpg.add_input_text(parent ='ADD_ROW',width = 400, label = GR_PROJ_label_list_ru[i], callback = add_row_text_input, enabled  = False, on_enter=True)
            elif i == 5:
                result = connection.execute(text("SELECT DISTINCT z2 FROM vuz" ))
                add_mas = []
                court_to_string(add_mas,result)
                add_mas.sort()
                add_row_id_1 = dpg.add_combo(parent ='ADD_ROW',width = 400, items = add_mas, label = GR_PROJ_label_list_ru[i], callback = add_row_text_input, enabled  = False)
                add_mas.clear()
            elif i == 6:
                add_row_id_1 = dpg.add_input_int(parent ='ADD_ROW',width = 400, label = GR_PROJ_label_list_ru[i], callback = add_row_text_input, min_value = 0,
                min_clamped = True, max_value = 999999999, max_clamped = True, enabled=False)
            elif i == 13:
                add_row_id_1 = dpg.add_input_text(parent ='ADD_ROW',width = 400, label = GR_PROJ_label_list_ru[i], multiline= True, callback = add_row_text_input, enabled  = False)
            else: 
                add_row_id_1 = dpg.add_input_text(parent ='ADD_ROW',width = 400, label = GR_PROJ_label_list_ru[i], callback = add_row_text_input, enabled  = False)
        add_row_id.append(add_row_id_1)

def edit_text_input(sender,specs):
    sender_label = dpg.get_item_label(sender)
    if sender_label == 'Конкурс':
        result = re.search(r'' + '\d\d' , specs)
        result = result.group(0)
        sender_label = 'Код конкурса'
        specs = result 
        result = connection.execute(text("SELECT COALESCE(MAX(g1)+1,1) FROM gr_proj WHERE codkon='" + str(specs) + "'"))
        add_mas = []
        court_to_string(add_mas,result)
        dpg.set_value(edit_id_input[1], add_mas[0])
        high_list[1][0] = str(specs)
        high_list[1][1] = add_mas[0]
        add_mas.clear()
        print('FROM EDIT INPUT')
        print(high_list)
        print(high_list_copy)
        return
    elif sender_label == 'Код НИР':
        result = re.search(r'' + '\d\d' , high_list[1][0])
        if result is not None:
            result = result.group(0)
            high_list[1][0] = result
        result = connection.execute(text("SELECT DISTINCT g1 FROM gr_proj WHERE codkon='" + high_list[1][0] + "'"))
        if result is None:
            high_list[1][1] = specs
        else:
            add_mas = []
            court_to_string(add_mas,result)
            if specs in add_mas:
                dpg.set_value(sender,high_list[1][1])
            else:
                high_list[1][1] = specs
        print(high_list)
        return
    elif sender_label == 'ГРНТИ':
        print(specs)
        result = re.search(r'' + '(\d\d\.\d\d(\.\d\d)?)(\,\d\d\.\d\d(\.\d\d)?)?' , specs)
        if result is None:
            dpg.set_value(sender,high_list_copy[3])
            return
        result = result.group(0)
        if len(result) != len(specs):
            print(specs + ' deleted2')
            dpg.set_value(sender,high_list_copy[3])
        else:
            high_list[1][3] = specs
        return
    elif sender_label == 'Руководитель НИР':
        high_list[1][2] = specs
        print(high_list)
        return
    elif sender_label == 'ВУЗ':
        high_list[1][4] = specs
        high_list[1][11] = str(VUZ_z2_codvuz_dict[specs])
        print(high_list)
        return
    elif sender_label == 'Плановый объем гранта':
        high_list[1][5] = str(specs)
        print(high_list)
        return
    elif sender_label == 'НИР':
        high_list[1][12] = specs
        print(high_list)
        return
    elif sender_label == 'Должность':
        high_list[1][13] = specs
        print(high_list)
        return
    elif sender_label == 'Ученое звание':
        high_list[1][14] = specs
        print(high_list)
        return
    else:
        high_list[1][15] = specs
        print(high_list)
        return

def edit_row_window_input_creation():
    for i in range(17):
        if i in [0,7,8,9,10,11,12]:
            continue
        else:
            if i == 1:
                dpg.add_combo(width = 400, parent ='EDIT',label = 'Конкурс', default_value = high_list[1][i-1],items = codkon_names_list,callback = edit_text_input)
            elif i == 2:
                dpg.add_input_text(parent ='EDIT',width = 400, label = GR_PROJ_label_list_ru[i], callback = edit_text_input, default_value = high_list[1][i-1], on_enter = True, decimal = True)
            elif i == 4:
                dpg.add_input_text(parent ='EDIT',width = 400, label = GR_PROJ_label_list_ru[i], callback = edit_text_input, default_value = high_list[1][i-1], on_enter = True)
            elif i == 5:
                result = connection.execute(text("SELECT DISTINCT z2 FROM vuz" ))
                add_mas = []
                court_to_string(add_mas,result)
                add_mas.sort()
                dpg.add_combo(width = 400, parent ='EDIT',label = GR_PROJ_label_list_ru[i], default_value = high_list[1][i-1],items = add_mas,callback = edit_text_input)
                add_mas.clear()
            elif i == 6:
                add_row_id_1 = dpg.add_input_int(parent ='EDIT',width = 400, label = GR_PROJ_label_list_ru[i], default_value =int(high_list[1][i-1]) ,callback = edit_text_input, min_value = 0,
                min_clamped = True, max_value = 999999999, max_clamped = True)
            elif i == 13:
                dpg.add_input_text(width = 400, parent ='EDIT',label = GR_PROJ_label_list_ru[i], default_value = high_list[1][i-1], callback = edit_text_input, multiline=True)
            else:
                dpg.add_input_text(width = 400, parent ='EDIT',label = GR_PROJ_label_list_ru[i], default_value = high_list[1][i-1], callback = edit_text_input)
            edit_id_input.append(dpg.last_item())

def gr_proj_table_recreation():
    dpg.delete_item(table_name_id[1][0], children_only=True)
    exec("for _label in " + "GR_PROJ" + "_label_list_ru: dpg.add_table_column(parent = table_name_id[1][0],label=_label)")
    result = connection.execute(text("SELECT codkon, g1,g8,g7,z2,g5,g2,g21,g22,g23,g24,codvuz,g6,g9,g10,g11 FROM gr_proj ORDER BY codkon ASC"))
    for _row in result:
        with dpg.table_row(parent = table_name_id[1][0]):
            for j in range(-1,len(_row)):
                if (j==-1):
                    dpg.add_selectable(height = 27, width = 27, callback = highlight_rows, user_data = 18)
                else:
                    dpg.add_text(f"{_row[j]}")

def gr_konk_table_recreation():
    dpg.delete_item(table_name_id[0][0], children_only=True)
    exec("for _label in " + "GR_KONK" + "_label_list_ru: dpg.add_table_column(parent = table_name_id[0][0],label=_label)")
    result = connection.execute(text("SELECT k2,codkon,npr,k12,k4,k41,k42,k43,k44 FROM gr_konk"))
    for row in result:
        with dpg.table_row(parent = table_name_id[0][0]):
            for j in range(-1,9):
                if (j==-1):
                    dpg.add_selectable(height = 27, width = 27, callback = highlight_rows, user_data = 11)
                else:
                    dpg.add_text(f"{row[j]}")

def info_selection(sender):
    if(dpg.get_item_label(sender) == 'gr_konk'):
        info = table_list_name_size[0]
    elif(dpg.get_item_label(sender) == 'gr_proj'):
        info = table_list_name_size[1]
    else:
        info = table_list_name_size[2]
    return info

def edit_callback():
    result = re.search(r'' + '\d\d' , high_list[1][0])
    if result is not None:
        result = result.group(0)
        high_list[1][0] = result
    
    edit_string = ''
    counter = 0
    for i in range(16):
        if high_list_copy[i] != high_list[1][i]:
            counter += 1

    if counter == 0:
        print('Нет изменений')
        return
    
    for i in range(17):
        if i == 0:
            continue
        else:
            if high_list_copy[i-1] != 'None' and i not in [7,8,9,10,11]:
                    edit_string += 'in_' + GR_PROJ_label_list[i] + ":='" + high_list_copy[i-1] + "',"
            else: 
                continue

    for i in range(16):
        if high_list_copy[i] != high_list[1][i] and counter != 1:    
            edit_string += 'out_' + GR_PROJ_label_list[i+1] + ":='" + high_list[1][i] + "',"
            counter -= 1
        elif high_list_copy[i] != high_list[1][i] and counter == 1:
            edit_string += 'out_' + GR_PROJ_label_list[i+1] + ":='" + high_list[1][i] + "'"
            
    try:
        connection.execute(text("SELECT edit_row_gr_proj("+ edit_string +")"))
        dpg.unhighlight_table_row(table = high_list[0][2], row = high_list[0][1])
        dpg.set_value(high_list[0][0],False)
        high_list.clear()
    except:
        print("Нельзя выполнить SQL запрос")
        for i in range(16):
            high_list[1][i] = high_list_copy[i]
        edit_btn()
        return
    connection.execute(text("COMMIT"))
    dpg.hide_item('EDIT')
    high_list_copy.clear()
    gr_proj_table_recreation()
    gr_konk_table_recreation()
    dpg.delete_item('F_S')
    global filter_condition
    global new_formed_filter_string
    filter_condition = False
    new_formed_filter_string = ''
    filter_condition_window()


def delete_callback():
    result = re.search(r'' + '\d\d' , high_list[1][0])
    if result is not None:
        result = result.group(0)
        high_list[1][0] = result
    delete_string = ''
    delete_string += "in_codkon:='" + high_list[1][0] + "',"
    delete_string += "in_g1:='" + high_list[1][1] + "'"
    print(delete_string)
    try:
        connection.execute(text("SELECT delete_row_gr_proj(" + delete_string + ")"))
        dpg.hide_item('EDIT')
        result = connection.execute(text("SELECT gr_konk_actual_financing()"))
        high_list.clear()
        high_list_copy.clear()
    except:
        print("Неверный SQL запрос")
        return
    connection.execute(text("COMMIT"))
    gr_proj_table_recreation()
    gr_konk_table_recreation()
    
    dpg.hide_item("DELETE")
    dpg.delete_item('F_S')
    global filter_condition
    global new_formed_filter_string
    filter_condition = False
    new_formed_filter_string = ''
    filter_condition_window()

def add_row_callback():
    if len(add_row_input)<2:
        return
    add_row_string = ''
    for i in range(len(add_row_input)):
        if i!=len(add_row_input)-1:
            add_row_string += 'in_' + add_row_input[i][1] +":='" + add_row_input[i][2] + "'" + ', '
        else:
            add_row_string += 'in_' + add_row_input[i][1] + ":='" + add_row_input[i][2] + "'"
    print(add_row_string)
    try:
        connection.execute(text("SELECT insert_into_gr_proj(" + add_row_string + ");"))
    except:
        print('Неверный SQL запрос')
        return
    add_row_input.clear()
    connection.execute(text('COMMIT'))
    hide_add_row_input()
    gr_proj_table_recreation()
    gr_konk_table_recreation()
    dpg.delete_item('F_S')
    global filter_condition
    global new_formed_filter_string
    filter_condition = False
    new_formed_filter_string = ''
    filter_condition_window()

def clear_add_row():
    for i in range(len(add_row_id)):
        dpg.delete_item(add_row_id[i])
    add_row_id.clear()
    add_grnti.clear()
    add_row_window_input_creation()

def delete_no_callback(sender):
    _window = dpg.get_item_parent(sender)
    dpg.hide_item(_window)

def table_creation(info):
    if info[0] == 'gr_proj':
        result = connection.execute(text("SELECT codkon, g1,g8,g7,z2,g5,g2,g21,g22,g23,g24,codvuz,g6,g9,g10,g11 FROM " + info[0] + " ORDER BY codkon ASC"))
    elif info[0] == 'gr_konk':
        result = connection.execute(text("SELECT k2,codkon,npr,k12,k4,k41,k42,k43,k44 FROM " + info[0]))
    else:
        result = connection.execute(text("SELECT * FROM " + info[0]))
    for row in result:
        with dpg.table_row():
            for j in range(-1,info[1]):
                if (j==-1):
                    dpg.add_selectable(height = 27, width = 27, callback = highlight_rows, user_data = info[1]+2)
                else:
                    dpg.add_text(f"{row[j]}")

def highlight_rows(sender, specs, user_data):
    _row = dpg.get_item_parent(sender)
    _table = dpg.get_item_parent(_row)
    first_row = dpg.get_item_children(_table, 1)[0]
    _row_ = int((_row-first_row)/user_data)
    cols = []
    for a in range(2,user_data):
        cols.append(dpg.get_value(_row+a)) 
    for i in codkon_names_list:
        result = re.search(r'' + cols[0] , i)
        if result is None:
            continue
        else:
            cols[0] = i
    if(specs):
        if(len(high_list)==0):
            dpg.highlight_table_row(table = _table, row = _row_, color = (0,0,100))
            high_list.append([sender,_row_,_table])
            high_list.append(cols)
        else:
            dpg.unhighlight_table_row(table = high_list[0][2], row = high_list[0][1])
            dpg.set_value(high_list[0][0],False)
            high_list.clear()
            high_list_copy.clear()
            dpg.highlight_table_row(table = _table, row = _row_, color = (0,0,100))
            high_list.append([sender,_row_,_table])
            high_list.append(cols)
    else:
        dpg.unhighlight_table_row(table = _table, row = _row_)
        high_list.clear()
        high_list_copy.clear()
    if len(high_list) != 0:
        if len(high_list_copy) != 0:
            high_list_copy.clear()
        for i in range(len(high_list[1])):
            if i == 0:
                result = re.search(r'' + '\d\d' , high_list[1][i])
                result = result.group(0)
                high_list_copy.append(result)
            else:
                high_list_copy.append(high_list[1][i])
    print('FROM HIGHLIGHT')
    print(high_list)
    print(high_list_copy)
    edit_btn()

def sort_algo(col_id, s_dir, row, sender):
    info = info_selection(sender)
    col_id = [x-row for x in col_id]
    sorting_string = ""
    counter = 1
    for id,dir in zip(col_id, s_dir):
        print(id, dir)
        sorting_string += eval(dpg.get_item_label(sender).upper() + "_label_list[id]")
        if(dir == 1):
            sorting_string += " ASC"
        else:
            sorting_string += " DESC"
        if (len(s_dir)-counter>0): 
            sorting_string += ", "
        counter +=1

    if(dpg.get_item_label(sender) == 'gr_proj'):
        result = connection.execute(text("SELECT codkon,g1,g8,g7,z2,g5,g2,g21,g22,g23,g24,codvuz,g6,g9,g10,g11 FROM " + dpg.get_item_label(sender) + " ORDER BY " + sorting_string))
    elif dpg.get_item_label(sender) == 'gr_konk':
        result = connection.execute(text("SELECT k2,codkon,npr,k12,k4,k41,k42,k43,k44 FROM " + dpg.get_item_label(sender) + " ORDER BY " + sorting_string))
    else:
        result = connection.execute(text("SELECT * FROM " + dpg.get_item_label(sender) + " ORDER BY " + sorting_string))
    dpg.delete_item(sender, children_only=True)
    exec("for _label in " + dpg.get_item_label(sender).upper() + "_label_list_ru: dpg.add_table_column(parent = sender,label=_label)")
    for _row in result:
        with dpg.table_row(parent = sender):
            for j in range(-1,len(_row)):
                if (j==-1):
                    dpg.add_selectable(height = 27, width = 27, callback = highlight_rows, user_data = info[1]+2)
                else:
                    dpg.add_text(f"{_row[j]}")
    high_list.clear()

def sort_callback(sender, sort_specs):
    print(sort_specs)
    global filter_condition
    global new_formed_filter_string
    info = info_selection(sender)
    if sort_specs is None: return
    if len(dpg.get_item_label(sort_specs[0][0]))==1: 
        for i in table_name_id:
            if i[1] == dpg.get_item_label(sender):
                dpg.delete_item(i[0], children_only=True)
                exec("for _label in " + dpg.get_item_label(sender).upper() + "_label_list_ru: dpg.add_table_column(parent = i[0],label=_label)")
                if dpg.get_item_label(sender) == 'gr_proj':
                    result = connection.execute(text( "SELECT codkon, g1,g8,g7,z2,g5,g2,g21,g22,g23,g24,codvuz,g6,g9,g10,g11 FROM " + info[0] + " ORDER BY codkon ASC"))
                elif dpg.get_item_label(sender) == 'gr_konk':
                    result = connection.execute(text( "SELECT k2,codkon,npr,k12,k4,k41,k42,k43,k44 FROM " + info[0]))
                else:
                    result = connection.execute(text("SELECT * FROM " + info[0]))
                for _row in result:
                    with dpg.table_row(parent = i[0]):
                        for j in range(-1,len(_row)):
                            if (j==-1):
                                dpg.add_selectable(height = 27, width = 27, callback = highlight_rows, user_data = info[1]+2)
                            else:
                                dpg.add_text(f"{_row[j]}")
        dpg.delete_item('F_S')
        filter_condition = False
        new_formed_filter_string = ''
        filter_condition_window()
        return
    
    if len(high_list)==1:
        dpg.unhighlight_table_row(table = high_list[0][2], row = high_list[0][1])
        dpg.set_value(high_list[0][0],False)
        high_list.clear()
    
    column_id = []
    direction = []
    if(len(sort_specs)==1):
        column_id.append(sort_specs[0][0])
        direction.append(sort_specs[0][1])
    else:
        for i in sort_specs:
            column_id.append(i[0])
            direction.append(i[1])

    row = dpg.get_item_children(sender, 1)[0]
    if(dpg.get_item_label(sender) == "gr_proj"):
        row-=17
    elif(dpg.get_item_label(sender) == 'vuz'):
        row-=12
    elif(dpg.get_item_label(sender) =='gr_konk'):
        row-=10

    sort_algo(column_id, direction, row, sender)
    res_mas.clear()
    column_id.clear()
    direction.clear()
    dpg.delete_item('F_S')
    filter_condition = False
    new_formed_filter_string = ''
    filter_condition_window()

def gr_konk_viz(sender):
    dpg.show_item('GR_KONK')

def gr_proj_viz(sender):
    dpg.show_item('GR_PROJ')

def vuz_viz(sender):
    dpg.show_item('VUZ')

def hide_filter_callback():
    if dpg.does_item_exist('FILTER'):
        dpg.hide_item('FILTER')
        dpg.delete_item('FILTER')
    if len(combo_id_mas) != 0:
        combo_id_mas.clear()

def hide_add_row_input():
    for i in add_row_id:
        dpg.delete_item(i)
    add_row_id.clear()
    add_row_window_input_creation()
    dpg.hide_item('ADD_ROW')

def reset_filter():
    for i in combo_id_mas:
        dpg.delete_item(i[1], children_only=False)

    specs_filter_list.clear()
    combo_id_mas.clear()

    if len(high_list)==1:
        dpg.unhighlight_table_row(table = high_list[0][2], row = high_list[0][1])
        dpg.set_value(high_list[0][0],False)
        high_list.clear()

    for i,j in {'region':'Фед. округ','Oblname':'Субъект фед.','City': 'Город','z1':'ВУЗ'}.items():
        result = connection.execute(text("""SELECT DISTINCT """ + i + """ FROM vuz vz
                                            INNER JOIN gr_proj gp ON vz.codvuz=gp.codvuz 
                                            INNER JOIN gr_konk gk ON gk.codkon = gp.codkon
                                         """))
        court_to_string(filter_mas_1,result)
        filter_mas_1.sort()
        combo = dpg.add_combo(parent = 'FILTER', label = j, callback = filtering, items = filter_mas_1, user_data = [i,filter_string])
        combo_id_mas.append([i,combo])
        filter_mas_1.clear()
    result = connection.execute(text("""SELECT DISTINCT k2 FROM gr_konk gk
                                        INNER JOIN gr_proj gp ON gk.codkon = gp.codkon
                                        INNER JOIN vuz vz ON vz.codvuz=gp.codvuz
                                     """))
    court_to_string(filter_mas_1,result)
    filter_mas_1.sort()
    combo = dpg.add_combo(parent = 'FILTER', label = 'Конкурс', callback = filtering, items = filter_mas_1, user_data = ['k2',filter_string])
    combo_id_mas.append(['k2',combo])
    filter_mas_1.clear()
    dpg.delete_item("F_S")
    filter_condition_window()

    global filter_condition
    filter_condition = False
    gr_proj_table_recreation()

def set_filter_callback(sender):
    if len(new_formed_filter_string)==0:
        return
    if dpg.does_item_exist('EDIT'):
        if dpg.is_item_shown('EDIT'):
            hide_edit()
    
    dpg.delete_item(table_name_id[1][0], children_only=True)
    exec("for _label in " + "GR_PROJ" + "_label_list_ru: dpg.add_table_column(parent = table_name_id[1][0],label=_label)")
    result = connection.execute(text("""SELECT gp.codkon,g1,g8,g7,gp.z2,g5,g2,g21,g22,g23,g24,gp.codvuz,g6,g9,g10,g11 
                                        FROM gr_proj gp
                                        INNER JOIN gr_konk gk ON gk.codkon = gp.codkon
                                        INNER JOIN vuz vz ON vz.codvuz=gp.codvuz
                                        WHERE """ + new_formed_filter_string))
    global filter_condition
    filter_condition = True
    for _row in result:
        with dpg.table_row(parent = table_name_id[1][0]):
            for j in range(-1,len(_row)):
                if (j==-1):
                    dpg.add_selectable(height = 27, width = 27, callback = highlight_rows, user_data = 18)
                else:
                    dpg.add_text(f"{_row[j]}")
    hide_filter_callback()
    dpg.delete_item("F_S")
    filter_condition_window()

def filtering(sender,specs, user_data):
    print(specs, dpg.get_item_label(sender))

    if len(high_list)==1:
        dpg.unhighlight_table_row(table = high_list[0][2], row = high_list[0][1])
        dpg.set_value(high_list[0][0],False)
        high_list.clear()

    cheker = False
    specs = to_cyr(specs)
    if [user_data[0],specs] not in specs_filter_list:
            if(len(specs_filter_list)==0):
                specs_filter_list.append([user_data[0], specs])
            else:
                for i in range(len(specs_filter_list)):
                    if(user_data[0] == specs_filter_list[i][0] and specs != specs_filter_list[i][1]):
                        specs_filter_list[i][1] = specs
                        cheker = True
                        break
                if(not cheker):
                    specs_filter_list.append([user_data[0], specs])
    print(specs_filter_list)
    
    if(user_data[1].find(user_data[0]) == -1):
        counter = 1
        for i in specs_filter_list:
            user_data[1]  += i[0] + '=' + "'"+ i[1] + "' "
            if len(specs_filter_list) - counter != 0:
                user_data[1] += ' AND '
                counter+=1
    else:
        result = re.search(r'' + user_data[0] + '.*\s' , user_data[1])
        result = result.group(0)
        user_data[1] = user_data[1].replace(result,'')
        counter = 1
        for i in specs_filter_list:
            user_data[1]  += i[0] + '=' + "'"+ i[1] + "' "
            if len(specs_filter_list) - counter != 0:
                    user_data[1] += ' AND '
                    counter+=1

    to_remove = []
    a=0
    for i in range(len(combo_id_mas)):
        for j in range(len(specs_filter_list)):
            if specs_filter_list[j][0] != combo_id_mas[i][0]:
                a+=1
            if a == len(specs_filter_list):
                dpg.delete_item(combo_id_mas[i][1])
                to_remove.append(combo_id_mas[i])
        a=0
    
    for i in to_remove:
        combo_id_mas.remove(i)
    to_remove.clear()

    check_mas = ['region','Oblname','City','z1','k2']
    for i in specs_filter_list:
        if i[0] in check_mas:
            check_mas.remove(i[0])

    for i in check_mas:
        if i != 'k2':
            result = connection.execute(text("""SELECT DISTINCT """ + i + """ FROM vuz vz
                                                INNER JOIN gr_proj gp ON vz.codvuz=gp.codvuz 
                                                INNER JOIN gr_konk gk ON gk.codkon = gp.codkon
                                                WHERE """ + user_data[1]))
            court_to_string(filter_mas_1,result)
            for c,j in {'region': 'Фед. округ' ,'Oblname': "Субъект фед.",'City': "Город" ,'z1': "ВУЗ" }.items():
                if c == i:
                    combo = dpg.add_combo(parent = 'FILTER', label = j , callback = filtering, user_data = [i,filter_string], items = filter_mas_1)
                    combo_id_mas.append([i,combo])
        else:
            result = connection.execute(text("""SELECT DISTINCT k2 FROM gr_konk gk
                                                INNER JOIN gr_proj gp ON gk.codkon = gp.codkon
                                                INNER JOIN vuz vz ON vz.codvuz=gp.codvuz
                                                WHERE """ + user_data[1] 
                                            ))
            court_to_string(filter_mas_1,result)
            combo = dpg.add_combo(parent = 'FILTER', label = "Конкурс" , callback = filtering, user_data = ['k2',filter_string], items = filter_mas_1)
            combo_id_mas.append(['k2',combo])
        filter_mas_1.clear()
    

    global new_formed_filter_string
    new_formed_filter_string = user_data[1]
    print(new_formed_filter_string)

def filter_btn(sender):
    specs_filter_list.clear()
    if(not dpg.does_item_exist('FILTER')):
        with dpg.window(tag = 'FILTER', label = "Фильтр", width = 520, height = 500, no_resize = True):
            for i,j in {'region':'Фед. округ','Oblname':'Субъект фед.','City': 'Город','z1':'ВУЗ'}.items():
                result = connection.execute(text("""SELECT DISTINCT """ + i + """ FROM vuz vz
                                                    INNER JOIN gr_proj gp ON vz.codvuz=gp.codvuz 
                                                    INNER JOIN gr_konk gk ON gk.codkon = gp.codkon
                                                 """))
                court_to_string(filter_mas_1,result)
                filter_mas_1.sort()
                combo = dpg.add_combo(label = j, callback = filtering, items = filter_mas_1, user_data = [i,filter_string])
                combo_id_mas.append([i,combo])
                filter_mas_1.clear()
            result = connection.execute(text("""SELECT DISTINCT k2 FROM gr_konk gk
                                                INNER JOIN gr_proj gp ON gk.codkon = gp.codkon
                                                INNER JOIN vuz vz ON vz.codvuz=gp.codvuz
                                             """))
            court_to_string(filter_mas_1,result)
            combo = dpg.add_combo(label = 'Конкурс', callback = filtering, items = filter_mas_1, user_data = ['k2',filter_string])
            combo_id_mas.append(['k2',combo])
            filter_mas_1.clear()
            with dpg.child_window(pos = (8,400),autosize_x=True, autosize_y=False):
                dpg.add_button(label = 'Установить фильтр', pos = (10,30), callback = set_filter_callback)
                dpg.add_button(label = 'Сбросить фильтр', pos = (200,30), callback = reset_filter)
                dpg.add_button(label = 'Выход/Отмена', pos = (370,30), callback = hide_filter_callback)
    else:                   
        if dpg.is_item_shown('FILTER'):
            return
        else:
            dpg.show_item('FILTER')
    
def delete_btn(sender):
    if len(high_list)==0 or dpg.get_item_label(high_list[0][2])!='gr_proj':
        return
    if(not dpg.does_item_exist('DELETE')):
        with dpg.window(tag = 'DELETE', label = "Удаление", width = 300, height = 200, modal = True, pos = (35,50)):
            if len(high_list)!=0:
                dpg.add_text("Удалить выделенную строку?")
                dpg.add_button(label = "Да", callback = delete_callback)
                dpg.add_button(label = "Нет", callback = delete_no_callback)
                dpg.add_button(label = "Отмена", callback = delete_no_callback)
            else:
                dpg.add_text("Выберите строку для удаления")
    else:
        if dpg.is_item_shown('DELETE'):
            return
        else:
            dpg.show_item('DELETE')

def edit_btn():
    if(not dpg.does_item_exist('EDIT')):
        if len(high_list)==0 or dpg.get_item_label(high_list[0][2])!='gr_proj':
            return
        else:
            with dpg.window(tag = 'EDIT', label = "Редактирование", width = 800, height = 600, pos = (30,250)):
                edit_row_window_input_creation()
                with dpg.child_window(pos = (8,500),autosize_x=True, autosize_y=False):
                    dpg.add_button(label = "Редактировать строку", callback = edit_callback, pos =(78,20), height = 50)
                    dpg.add_button(label = "Сброс изменений", callback = clear_edit_row, pos = (368,20),  height = 50)
                    dpg.add_button(label = "Выход", callback = hide_edit, pos = (608,20), height = 50 )
    else:
        if len(high_list)==0 or dpg.get_item_label(high_list[0][2])!='gr_proj':
            return
        else:
            for x in edit_id_input:
                dpg.delete_item(x)
            edit_id_input.clear()
            if(len(high_list)):
                edit_row_window_input_creation()
        dpg.show_item('EDIT')

def add_row_text_input(sender, specs):
    sender_label = dpg.get_item_label(sender)
    if sender_label == 'Плановый объем гранта':
        specs = str(specs)

    specs = to_cyr(specs)
    global add_row_max_g1
    global add_row_input

    if sender_label == 'Код НИР':
        for i in add_row_input:
            if i[0] == 2:
                add_row_input.remove(i)
                break
        result = connection.execute(text("SELECT DISTINCT g1 FROM gr_proj WHERE codkon='" + add_row_input[0][2] + "'"))
        if result is None:
            if int(specs) >= 1:
                add_row_input.append([2,GR_PROJ_label_list[2],specs])
            else:
                dpg.set_value(sender,'')
        else:
            add_mas = []
            court_to_string(add_mas,result)
            if specs in add_mas:
                dpg.set_value(sender,'')
            else:
                if int(specs) >= 1:
                    add_row_input.append([2,GR_PROJ_label_list[2],specs])
                else:
                    dpg.set_value(sender,'')
        print(add_row_input)
        return

    if sender_label == 'Конкурс':
        result = re.search(r'' + '\d\d' , specs)
        result = result.group(0)
        sender_label = 'Код конкурса'
        specs = result 

    if sender_label == 'ГРНТИ':
        print(specs)
        result = re.search(r'' + '(\d\d\.\d\d(\.\d\d)?)(\,\d\d\.\d\d(\.\d\d)?)?' , specs)
        if result is None:
           for i in add_row_input:
                if i[0] == 4:
                    add_row_input.remove(i)
                    break
                dpg.set_value(sender,'')
                return
        result = result.group(0)
        checker = False
        if len(result) != len(specs):
            print(specs + ' deleted2')
            for i in add_row_input:
                if i[0] == 4:
                    add_row_input.remove(i)
                    break
            dpg.set_value(sender,'')
        else:
            for i in range(len(add_row_input)):
                if add_row_input[i][0] == 4:
                    add_row_input[i] = [4,GR_PROJ_label_list[4],specs]
                    checker = True
                    break
            if not checker:
                add_row_input.append([4,GR_PROJ_label_list[4],specs])
            print(add_row_input)
        return

    for i in range(len(GR_PROJ_label_list_ru)):
        if i == 0:
            continue
        if sender_label == GR_PROJ_label_list_ru[i]:
            for j in range(len(add_row_input)):
                if i == add_row_input[j][0]:
                    add_row_input[j][2] = specs
            if [i,GR_PROJ_label_list[i],specs] not in add_row_input:
                if sender_label == 'ВУЗ':
                    add_row_input.append([i,GR_PROJ_label_list[i],specs])
                    add_row_input.append([12,GR_PROJ_label_list[12],str(VUZ_z2_codvuz_dict[specs])])
                else:
                    add_row_input.append([i,GR_PROJ_label_list[i],specs])
    
    if add_row_input[0][1] == 'codkon' and sender_label == 'Код конкурса':
        add_row_choose_codkon = add_row_input[0][2]
        result = connection.execute(text("SELECT COALESCE(MAX(g1)+1,1) FROM gr_proj WHERE codkon='" + str(add_row_choose_codkon) + "'"))
        add_mas = []
        court_to_string(add_mas,result)
        add_row_max_g1 = add_mas[0]
        dpg.set_value(add_row_id[1], add_mas[0])
        if not dpg.is_item_enabled(add_row_id[1]):
            for i in range(1,len(add_row_id)):
                dpg.enable_item(add_row_id[i])
        add_mas.clear()

def clear_edit_row():
    for i in edit_id_input:
        dpg.delete_item(i)
    edit_id_input.clear()
    print(high_list_copy)
    for i in range(len(high_list_copy)):
        if i == 0:
            for j in codkon_names_list:
                result = re.search(r'' + high_list_copy[i], j)
                if result is None:
                    continue
                else:
                    high_list[1][i] = j
        else:
            high_list[1][i] = high_list_copy[i]
    edit_row_window_input_creation()

def hide_edit():
    dpg.unhighlight_table_row(table = high_list[0][2], row = high_list[0][1])
    dpg.set_value(high_list[0][0],False)
    high_list.clear()
    high_list_copy.clear()
    dpg.hide_item('EDIT')
        
def add_row_btn(sender):
    if(not dpg.does_item_exist('ADD_ROW')):
        with dpg.window(tag = 'ADD_ROW', label = "Добавить строку", width = 800, height = 600, pos = (35,50)):
            add_row_window_input_creation()
            with dpg.child_window(pos = (8,500),autosize_x=True, autosize_y=False):
                dpg.add_button(label = "Добавить строку", callback = add_row_callback, pos =(128,20), height = 50)
                dpg.add_button(label = "Сброс ввода", callback = clear_add_row, pos = (368,20),  height = 50)
                dpg.add_button(label = "Выход", callback = hide_add_row_input, pos = (608,20), height = 50 )
    dpg.show_item('ADD_ROW')

def export_fed_callback():
    excel_doc = op.Workbook()
    excel_doc.create_sheet(title = 'Субъекты федерации', index = 0)
    sheetnames = excel_doc.sheetnames
    sheet = excel_doc[sheetnames[0]]
    #FED
    sheet[f'A{1}'] = '№' 
    sheet[f'B{1}'] = 'Название субъекта'
    sheet[f'C{1}'] = 'Кол-во НИР' 
    sheet[f'D{1}'] = 'Плановое финансирование'
    sheet[f'E{1}'] = 'Кол-во конкурсов в субъекте'
    sheet[f'F{1}'] = 'Кол-во ВУЗов в субъекте'

    i=2
    if filter_condition:
        filter_with_where = ' WHERE ' + new_formed_filter_string
        result = connection.execute(text(FED_analys_sql_string[:597] + filter_with_where + FED_analys_sql_string[597:]))
    else:
        result = connection.execute(text(FED_analys_sql_string))
    for row in result:
        a,b,c,d,e,f = row
        sheet[f'A{i}'] = a 
        sheet[f'B{i}'] = b 
        sheet[f'C{i}'] = c 
        sheet[f'D{i}'] = d 
        sheet[f'E{i}'] = e
        sheet[f'F{i}'] = f 
        i+=1

    #TOTAL FED
    if filter_condition:
        filter_with_where = ' WHERE ' + new_formed_filter_string
        result = connection.execute(text(TOTAL_for_fed_analys_sql_string[:434] + filter_with_where + TOTAL_for_fed_analys_sql_string[434:] + filter_with_where))
    else:
        result = connection.execute(text(TOTAL_for_fed_analys_sql_string))
    
    for row in result:
        a,b,c,d,e = row
        sheet[f'A{i}'] = a 
        sheet[f'B{i}'] = b 
        sheet[f'C{i}'] = c 
        sheet[f'D{i}'] = d 
        sheet[f'E{i}'] = e

    #FILTER 
    sheet[f'I{1}'] = 'Фильтр'
    i=2
    for a in specs_filter_list:
        if a[0] == 'Oblname':
            sheet[f'H{i}'] = 'Область'
        elif a[0] == 'region':
            sheet[f'H{i}'] = 'Регион'
        elif a[0] == 'City':
            sheet[f'H{i}'] = 'Город'
        elif a[0] == 'z1':
            sheet[f'H{i}'] = 'ВУЗ'
        else:
            sheet[f'H{i}'] = 'Конкурс'     
        sheet[f'I{i}'] = a[1]
        i+=1

    global fed_ex_name
    excel_doc.save('FED_ANALYS_'+ fed_ex_name +'.xlsx')
    
    with dpg.window(tag = 'EX_FED',label = "Формирование документа", width = 500, height = 200, modal = True, on_close = hide_ex_fed):
        dpg.add_text("Был сформирован документ 'FED_ANALYS_"+ fed_ex_name +".xlsx' ")
        fed_ex_name = str(int(fed_ex_name)+1)
        dpg.add_button(label = "Закрыть", callback = hide_ex_fed )

def hide_ex_fed():
    dpg.hide_item('EX_FED')
    dpg.delete_item('EX_FED',children_only = False)

def export_konk_callback():
    excel_doc = op.Workbook()
    excel_doc.create_sheet(title = 'Конкурсы', index = 0)
    sheetnames = excel_doc.sheetnames
    sheet = excel_doc[sheetnames[0]]
    #KONK
    sheet[f'A{1}'] = '№' 
    sheet[f'B{1}'] = 'Название конкурса'
    sheet[f'C{1}'] = 'Кол-во НИР' 
    sheet[f'D{1}'] = 'Плановое финансирование' 
    sheet[f'E{1}'] = 'Кол-во ВУЗов' 

    i=2
    if filter_condition:
        filter_with_where = ' WHERE ' + new_formed_filter_string
        result = connection.execute(text(KONK_analys_sql_string[:340] + filter_with_where + KONK_analys_sql_string[340:]))
    else:
        result = connection.execute(text(KONK_analys_sql_string))
    
    for row in result:
        a,b,c,d,e = row
        sheet[f'A{i}'] = a 
        sheet[f'B{i}'] = b 
        sheet[f'C{i}'] = c 
        sheet[f'D{i}'] = d 
        sheet[f'E{i}'] = e 
        i+=1

    #TOTAL KONK
    if filter_condition:
        filter_with_where = ' WHERE ' + new_formed_filter_string
        result = connection.execute(text(TOTAL_for_vuz_konk_analys_sql_string + filter_with_where))
    else:
        result = connection.execute(text(TOTAL_for_vuz_konk_analys_sql_string))
    
    for row in result:
        a,b,c,d,e = row
        sheet[f'A{i}'] = a 
        sheet[f'B{i}'] = b 
        sheet[f'C{i}'] = c 
        sheet[f'D{i}'] = d 
        sheet[f'E{i}'] = e

        #FILTER 
    sheet[f'I{1}'] = 'Фильтр'
    i=2
    for a in specs_filter_list:
        if a[0] == 'Oblname':
            sheet[f'H{i}'] = 'Область'
        elif a[0] == 'region':
            sheet[f'H{i}'] = 'Регион'
        elif a[0] == 'City':
            sheet[f'H{i}'] = 'Город'
        elif a[0] == 'z1':
            sheet[f'H{i}'] = 'ВУЗ'
        else:
            sheet[f'H{i}'] = 'Конкурс'     
        sheet[f'I{i}'] = a[1]
        i+=1
    
    global konk_ex_name
    excel_doc.save('KONK_ANALYS_'+ konk_ex_name +'.xlsx')

    with dpg.window(tag = 'EX_KONK',label = "Формирование документа", width = 500, height = 200, modal = True, on_close = hide_ex_konk ):
        dpg.add_text("Был сформирован документ 'KONK_ANALYS_" + konk_ex_name + ".xlsx' ")
        konk_ex_name = str(int(konk_ex_name)+1)
        dpg.add_button(label = "Закрыть", callback = hide_ex_konk)

def hide_ex_konk():
    dpg.hide_item('EX_KONK')
    dpg.delete_item('EX_KONK',children_only = False)

def export_vuz_callback():
    excel_doc = op.Workbook()
    excel_doc.create_sheet(title = 'ВУЗы', index = 0)
    sheetnames = excel_doc.sheetnames
    sheet = excel_doc[sheetnames[0]]
    #PROJ
    sheet[f'A{1}'] = '№' 
    sheet[f'B{1}'] = 'ВУЗ'
    sheet[f'C{1}'] = 'Кол-во НИР' 
    sheet[f'D{1}'] = 'Плановое финансирование' 
    sheet[f'E{1}'] = 'Кол-во конкурсов по грантам' 
    i=2
    global filter_condition
    global new_formed_filter_string
    
    if filter_condition:
        filter_with_where = ' WHERE ' + new_formed_filter_string
        result = connection.execute(text(VUZ_analys_sql_string[:339] + filter_with_where + VUZ_analys_sql_string[339:]))
    else:
        result = connection.execute(text(VUZ_analys_sql_string))

    for row in result:
        a,b,c,d,e = row
        sheet[f'A{i}'] = a 
        sheet[f'B{i}'] = b 
        sheet[f'C{i}'] = c 
        sheet[f'D{i}'] = d 
        sheet[f'E{i}'] = e 
        i+=1

    #TOTAL PROJ
    if filter_condition:
        filter_with_where = ' WHERE ' + new_formed_filter_string
        result = connection.execute(text(TOTAL_for_vuz_konk_analys_sql_string + filter_with_where))
    else:
        result = connection.execute(text(TOTAL_for_vuz_konk_analys_sql_string))

    for row in result:
        a,b,c,d,e = row
        sheet[f'A{i}'] = a 
        sheet[f'B{i}'] = b 
        sheet[f'C{i}'] = c 
        sheet[f'D{i}'] = d 
        sheet[f'E{i}'] = e 

    #FILTER 
    sheet[f'I{1}'] = 'Фильтр'
    i=2
    for a in specs_filter_list:
        if a[0] == 'Oblname':
            sheet[f'H{i}'] = 'Область'
        elif a[0] == 'region':
            sheet[f'H{i}'] = 'Регион'
        elif a[0] == 'City':
            sheet[f'H{i}'] = 'Город'
        elif a[0] == 'z1':
            sheet[f'H{i}'] = 'ВУЗ'
        else:
            sheet[f'H{i}'] = 'Конкурс'     
        sheet[f'I{i}'] = a[1]
        i+=1
    
    global vuz_ex_name
    excel_doc.save('VUZ_ANALYS_'+ vuz_ex_name +'.xlsx')   
    
    with dpg.window(tag = 'EX_VUZ',label = "Формирование документа", width = 500, height = 200, modal = True,on_close = hide_ex_vuz):
        dpg.add_text("Был сформирован документ 'VUZ_ANALYS_" + vuz_ex_name + ".xlsx' ")
        vuz_ex_name = str(int(vuz_ex_name)+1)
        dpg.add_button(label = "Закрыть", callback = hide_ex_vuz)

def hide_ex_vuz():
    dpg.hide_item('EX_VUZ')
    dpg.delete_item('EX_VUZ',children_only = False)

def hide_ex_viz_vuz():
    dpg.hide_item('EXPORT_VUZ')
    dpg.delete_item('EXPORT_VUZ')

def hide_ex_viz_konk():
    dpg.hide_item('EXPORT_KONK')
    dpg.delete_item('EXPORT_KONK')

def hide_ex_viz_fed():
    dpg.hide_item('EXPORT_FED')
    dpg.delete_item('EXPORT_FED')

def vuz_analys_sql_request(filter_string, condition):
    if condition:
        filter_with_where = ' WHERE ' + filter_string
        result = connection.execute(text(VUZ_analys_sql_string[:339] + filter_with_where + VUZ_analys_sql_string[339:]))
    else:
        result = connection.execute(text(VUZ_analys_sql_string))
    return result

def ex_vuz_creation():
    global new_formed_filter_string
    global filter_condition
    result = vuz_analys_sql_request(new_formed_filter_string,filter_condition)
    for row in result:
        with dpg.table_row():
            for j in range(len(row)):
                    dpg.add_text(f"{row[j]}")

def konk_analys_sql_request(filter_string, condition):
    if condition:
        filter_with_where = ' WHERE ' + filter_string
        result = connection.execute(text(KONK_analys_sql_string[:340] + filter_with_where + KONK_analys_sql_string[340:]))
    else:
        result = connection.execute(text(KONK_analys_sql_string))
    return result

def ex_konk_creation():
    global new_formed_filter_string
    global filter_condition
    result = konk_analys_sql_request(new_formed_filter_string,filter_condition)
    for row in result:
        with dpg.table_row():
            for j in range(len(row)):
                    dpg.add_text(f"{row[j]}")

def fed_analys_sql_request(filter_string, condition):
    if condition:
        filter_with_where = ' WHERE ' + filter_string
        result = connection.execute(text(FED_analys_sql_string[:597] + filter_with_where + FED_analys_sql_string[597:]))
    else:
        result = connection.execute(text(FED_analys_sql_string))
    return result

def ex_fed_creation():
    global new_formed_filter_string
    global filter_condition
    result = fed_analys_sql_request(new_formed_filter_string,filter_condition)
    for row in result:
        with dpg.table_row():
            for j in range(len(row)):
                    dpg.add_text(f"{row[j]}")

def TOTAL_for_VUZ_KONK_analys(filter_string, condition):
    if condition:
        filter_with_where = ' WHERE ' + filter_string
        result = connection.execute(text(TOTAL_for_vuz_konk_analys_sql_string + filter_with_where))
    else:
        result = connection.execute(text(TOTAL_for_vuz_konk_analys_sql_string))
    return result

def ex_for_vuz_konk_end():
    global new_formed_filter_string
    global filter_condition
    result = TOTAL_for_VUZ_KONK_analys(new_formed_filter_string,filter_condition)
    for row in result:
        with dpg.table_row():
            for j in range(len(row)):
                    dpg.add_text(f"{row[j]}")

def TOTAL_for_FED_analys(filter_string, condition):
    if condition:
        filter_with_where = ' WHERE ' + filter_string
        result = connection.execute(text(TOTAL_for_fed_analys_sql_string[:434] + filter_with_where + TOTAL_for_fed_analys_sql_string[434:] + filter_with_where))
    else:
        result = connection.execute(text(TOTAL_for_fed_analys_sql_string))
    return result

def ex_for_fed_end():
    global new_formed_filter_string
    global filter_condition
    result = TOTAL_for_FED_analys(new_formed_filter_string,filter_condition)
    for row in result:
        with dpg.table_row():
            for j in range(len(row)):
                    dpg.add_text(f"{row[j]}")

def ex_viz_vuz():
    if(not dpg.does_item_exist('EXPORT_VUZ')):
        with dpg.window(tag = 'EXPORT_VUZ', label = "Анализ ВУЗов", width = 900, height = 600, pos = (35,50)):
            with dpg.table(height = 450, label = 'По ВУЗам', header_row=True,
                        borders_innerH=True, borders_outerH=True, borders_innerV=True,
                        borders_outerV=True, resizable = False, scrollX=True, scrollY=True, no_keep_columns_visible=True, 
                        policy = dpg.mvTable_SizingFixedFit, sortable = False):
                dpg.add_table_column(label = '№')
                dpg.add_table_column(label = 'ВУЗ')
                dpg.add_table_column(label = 'Кол-во НИР')
                dpg.add_table_column(label = 'Плановое финансирование')
                dpg.add_table_column(label = 'Количество конкурсов грантов')
                ex_vuz_creation()
                ex_for_vuz_konk_end()
            with dpg.child_window(pos = (8,500),autosize_x=True, autosize_y=False):
                dpg.add_button(label = "Экспорт в файл", callback = export_vuz_callback, pos =(128,20), height = 50)
                dpg.add_button(label = "Выход", callback = hide_ex_viz_vuz, pos = (608,20), height = 50)

    dpg.show_item('EXPORT_VUZ')

def ex_viz_konk():
    if(not dpg.does_item_exist('EXPORT_KONK')):
        with dpg.window(tag = 'EXPORT_KONK', label = "Анализ конкурсов", width = 900, height = 600, pos = (35,50)):
            with dpg.table(height = 450, label = 'По конкурсам', header_row=True,
                    borders_innerH=True, borders_outerH=True, borders_innerV=True,
                    borders_outerV=True, resizable = False, scrollX=True, scrollY=True, no_keep_columns_visible=True, 
                    policy = dpg.mvTable_SizingFixedFit, sortable = False):
                dpg.add_table_column(label = '№')
                dpg.add_table_column(label = 'Название конкурса')
                dpg.add_table_column(label = 'Кол-во НИР')
                dpg.add_table_column(label = 'Плановое финансирование')
                dpg.add_table_column(label = 'Количество ВУЗов')
                ex_konk_creation()
                ex_for_vuz_konk_end()
            with dpg.child_window(pos = (8,500),autosize_x=True, autosize_y=False):
                dpg.add_button(label = "Экспорт в файл", callback = export_konk_callback, pos =(128,20), height = 50)
                dpg.add_button(label = "Выход", callback = hide_ex_viz_konk, pos = (608,20), height = 50)

    dpg.show_item('EXPORT_KONK')

def ex_viz_fed():
    if(not dpg.does_item_exist('EXPORT_FED')):
        with dpg.window(tag = 'EXPORT_FED', label = "Анализ субъектов федерации", width = 900, height = 600, pos = (35,50)):
            with dpg.table(height = 450, label = 'По субъектам федерации', header_row=True,
                    borders_innerH=True, borders_outerH=True, borders_innerV=True,
                    borders_outerV=True, resizable = False, scrollX=True, scrollY=True, no_keep_columns_visible=True, 
                    policy = dpg.mvTable_SizingFixedFit, sortable = False):
                dpg.add_table_column(label = '№')
                dpg.add_table_column(label = 'Название субъекта')
                dpg.add_table_column(label = 'Кол-во НИР')
                dpg.add_table_column(label = 'Плановое финансирование')
                dpg.add_table_column(label = 'Кол-во ВУЗов в субъекте')
                dpg.add_table_column(label = 'Кол-во конкурсов в субъекте')
                ex_fed_creation()
                ex_for_fed_end()
            with dpg.child_window(pos = (8,500),autosize_x=True, autosize_y=False):
                dpg.add_button(label = "Экспорт в файл", callback = export_fed_callback, pos =(128,20), height = 50)
                dpg.add_button(label = "Выход", callback = hide_ex_viz_fed, pos = (608,20), height = 50)

    dpg.show_item('EXPORT_FED')

def filter_condition_window():
    if not dpg.does_item_exist('F_S'):
        with dpg.child_window(parent = 'Main Window', tag = 'F_S', pos=(0,25),width = dpg.get_viewport_width(), height = 55,no_scrollbar = False, horizontal_scrollbar = True):
            global filter_condition
            if filter_condition:
                string = ''
                for i in specs_filter_list:
                    if i[0] == 'Oblname':
                        string += ' Область - '  + i[1] + ";"
                    elif i[0] == 'region':
                        string += ' Регион - '  + i[1] + ";"
                    elif i[0] == 'City':
                        string += ' Город - '  + i[1] + ";"
                    elif i[0] == 'z1':
                        string += ' ВУЗ - '  + i[1] + ";"
                    else:
                        string += ' Конкурс - '  + i[1] + ";"
                dpg.add_text(default_value = 'Состояние фильтра:' + string)
            else:
                dpg.add_text(default_value = 'Состояние фильтра:')

def finance_table_creation(sender,specs):
    if dpg.does_item_exist('FINANCE_TABLE'):
        dpg.delete_item('FINANCE_TABLE', children_only = True)
    result = re.search(r'' + '\d' , specs)
    if result is None:
        return
    specs = result.group(0)
    result = connection.execute(text("""SELECT ROW_NUMBER() OVER(ORDER BY gr.z2 ASC) ,gr.z2, SUM(gr.g2""" + specs + """)
                                        FROM gr_konk gk
                                        INNER JOIN gr_proj gr ON gr.codkon=gk.codkon
                                        INNER JOIN vuz vz ON gr.codvuz = vz.codvuz
                                        GROUP BY gr.z2"""))
    dpg.add_table_column(parent = 'FINANCE_TABLE',label = '№')
    dpg.add_table_column(parent = 'FINANCE_TABLE',label = 'ВУЗ')
    dpg.add_table_column(parent = 'FINANCE_TABLE',label = 'Фактическое финансирование')
    for row in result:
        with dpg.table_row(parent = 'FINANCE_TABLE'):
            for j in range(len(row)):
                    dpg.add_text(f"{row[j]}")
    result = connection.execute(text("""SELECT '', 'ИТОГО', SUM(gr.g2""" + specs + """)
                                        FROM gr_konk gk
                                        INNER JOIN gr_proj gr ON gr.codkon=gk.codkon
                                        INNER JOIN vuz vz ON gr.codvuz = vz.codvuz"""))
    for row in result:
        with dpg.table_row(parent = 'FINANCE_TABLE'):
            for j in range(len(row)):
                    dpg.add_text(f"{row[j]}")



def refresh_actual_financing():
    result = connection.execute(text("SELECT SUM(gk.k4) FROM gr_konk gk"))
    add_mas = []
    court_to_string(add_mas,result)
    global actual_financing
    actual_financing = int(add_mas[0])
    print('Фактическое финансирование: '+ str(actual_financing))
    add_mas.clear()

def refresh_planned_financing():
    result = connection.execute(text("SELECT SUM(gk.k12) FROM gr_konk gk"))
    add_mas = []
    court_to_string(add_mas,result)
    global planned_financing
    planned_financing = int(add_mas[0])
    print('Планируемое финансирование: '+ str(planned_financing))
    add_mas.clear()

def correct_procent(sender,specs):
    global planned_financing
    procent = 100*specs/planned_financing
    dpg.set_value(finance_id[2],procent)
    global current_int
    global current_procent
    current_int = specs
    current_procent = procent

def correct_int(sender,specs):
    global planned_financing
    value = specs*planned_financing//100
    dpg.set_value(finance_id[1],value)
    global current_int
    global current_procent
    current_int = value
    current_procent = specs

def finance_subwindow():
    finance_id.clear()
    refresh_planned_financing()
    refresh_actual_financing()
    global planned_financing
    global actual_financing
    if dpg.does_item_exist('FINANCE_SUB'):
        dpg.delete_item('FINANCE_SUB')
    with dpg.child_window(tag = 'FINANCE_SUB',parent ='FINANCE',pos = (8,34), height = 450, width = 640):
        dpg.add_text(default_value = 'Плановое финансирование ' + str(planned_financing))
        dpg.add_text(default_value = 'Выделенное финансирование ' + str(actual_financing) + ' (' + str(round(100*actual_financing/planned_financing,1)) + '% от планового)')
        dpg.add_combo(width = 250 ,items = finance_items,callback = finance_table_creation)
        finance_id.append(dpg.last_item())
        dpg.add_input_int(width = 250, label = 'руб', callback = correct_procent, min_value=0, min_clamped=True, max_value = planned_financing-actual_financing if planned_financing-actual_financing>0 else 0 , max_clamped = True)
        finance_id.append(dpg.last_item())
        dpg.add_input_float(width = 250, label = '%', callback = correct_int,  min_value=0, min_clamped=True, max_value = 100*((planned_financing-actual_financing)/planned_financing) if 100*((planned_financing-actual_financing)/planned_financing) > 0 else 0, max_clamped = True)
        finance_id.append(dpg.last_item())

def enter_finance_callback():
    result = dpg.get_value(finance_id[0])
    if result is None or result == '':
        return
    print(result)
    if result == '1 квартал':
        column_name = 'g21'
    elif result == '2 квартал':
        column_name = 'g22'
    elif result == '3 квартал':
        column_name = 'g23'
    else:
        column_name = 'g24'
    print(column_name)
    result = connection.execute(text("SELECT COUNT(g1) FROM gr_proj"))
    add_mas = []
    court_to_string(add_mas, result)
    nir_number = int(add_mas[0])
    global current_int
    if current_int == 0 or current_int is None:
        print('Введите значение')
        return
    portion = int(current_int // nir_number)
    print(portion)
    print(column_name)
    try:
        result = connection.execute(text("SELECT gr_proj_kv_financing('" + column_name + "'," + str(portion) + ")"))
        if column_name == 'g21':    
            result = connection.execute(text("SELECT kv1()"))
        elif column_name == 'g22':
            result = connection.execute(text("SELECT kv2()"))
        elif column_name == 'g23':
            result = connection.execute(text("SELECT kv3()"))
        else:
            result = connection.execute(text("SELECT kv4()"))
        result = connection.execute(text("SELECT gr_konk_actual_financing()"))
        result = connection.execute(text("SELECT gr_proj_actual_fin()"))
    except:
        print('Ошибка')
        return
    result = connection.execute(text("COMMIT"))
    gr_proj_table_recreation()
    gr_konk_table_recreation()
    finance_table_creation(finance_id[0],dpg.get_value(finance_id[0]))
    finance_subwindow()
    current_int = 0
    
def hide_ex_finance():
    dpg.hide_item('EX_FINANCE')
    dpg.delete_item('EX_FINANCE')

def export_finance_callback():
    kv = dpg.get_value(finance_id[0])
    if kv is None:
        return
    result = re.search(r'' + '\d' , kv)
    if result is None:
        return
    kv = result.group(0)
    excel_doc = op.Workbook()
    excel_doc.create_sheet(title = 'Финансирование', index = 0)
    sheetnames = excel_doc.sheetnames
    sheet = excel_doc[sheetnames[0]]

    #FINANCE
    sheet[f'A{1}'] = '№' 
    sheet[f'B{1}'] = 'ВУЗ'
    sheet[f'C{1}'] = 'Финансирование за квартал' 
    i=2
    result = connection.execute(text("""SELECT ROW_NUMBER() OVER(ORDER BY gr.z2 ASC) ,gr.z2, SUM(gr.g2""" + kv + """)
                                        FROM gr_konk gk
                                        INNER JOIN gr_proj gr ON gr.codkon=gk.codkon
                                        INNER JOIN vuz vz ON gr.codvuz = vz.codvuz
                                        GROUP BY gr.z2"""))

    for row in result:
        a,b,c = row
        sheet[f'A{i}'] = a 
        sheet[f'B{i}'] = b 
        sheet[f'C{i}'] = c 
        i+=1

    result = connection.execute(text("""SELECT '', 'ИТОГО', SUM(gr.g2""" + kv + """)
                                        FROM gr_konk gk
                                        INNER JOIN gr_proj gr ON gr.codkon=gk.codkon
                                        INNER JOIN vuz vz ON gr.codvuz = vz.codvuz"""))
    #TOTAL FINANCE
    for row in result:
        a,b,c = row
        sheet[f'A{i}'] = a 
        sheet[f'B{i}'] = b 
        sheet[f'C{i}'] = c 

    global finance_ex_name
    excel_doc.save('FINANCE_ANALYS_'+ finance_ex_name +'.xlsx')   
    
    with dpg.window(tag = 'EX_FINANCE',label = "Формирование документа", width = 500, height = 200, modal = True,on_close = hide_ex_vuz):
        dpg.add_text("Был сформирован документ 'FINANCE_ANALYS_" + finance_ex_name + ".xlsx' ")
        finance_ex_name= str(int(finance_ex_name)+1)
        dpg.add_button(label = "Закрыть", callback = hide_ex_finance)


def hide_finance():
    dpg.hide_item('FINANCE')
    dpg.delete_item('FINANCE')
    finance_id.clear()
    return

def finance_window():
    if dpg.does_item_exist("FINANCE"):
        dpg.show_item("FINANCE")
    else:
        with dpg.window(tag = "FINANCE", width = 1296, height = 600, label = 'Финансирование'):
            with dpg.child_window(parent ='FINANCE', pos = (648,34), width=1280):
                with dpg.table( tag = 'FINANCE_TABLE',height = 450, width = 632, label = 'Финансирование', header_row=True,
                                borders_innerH=True, borders_outerH=True, borders_innerV=True,
                                borders_outerV=True, resizable = False, scrollX=True, scrollY=True, no_keep_columns_visible=True, 
                                policy = dpg.mvTable_SizingFixedFit, sortable = False):
                    dpg.add_table_column(label = '№')
                    dpg.add_table_column(label = 'ВУЗ')
                    dpg.add_table_column(label = 'Фактическое финансирование')
            finance_subwindow()
            with dpg.child_window(parent ='FINANCE', pos = (8,490), width=1280):
                dpg.add_button(label = "Подтвердить ввод", callback = enter_finance_callback, pos =(128,20), height = 50)
                dpg.add_button(label = "Экспорт в файл", callback = export_finance_callback, pos =(608,20), height = 50)
                dpg.add_button(label = "Выход", callback = hide_finance, pos = (848,20), height = 50)



with dpg.window(tag = 'Main Window', height = 720, width = 1280):
    dpg.set_primary_window("Main Window", True)
    with dpg.viewport_menu_bar():
        with dpg.menu(label="Таблицы"):
            dpg.add_menu_item(label='Конкурсы' , callback = gr_konk_viz)
            dpg.add_menu_item(label='Проекты', callback = gr_proj_viz)
            dpg.add_menu_item(label='ВУЗы', callback = vuz_viz)
        with dpg.menu(label="Анализ"):
            dpg.add_menu_item(label="Анализ по ВУЗам", callback = ex_viz_vuz)
            dpg.add_menu_item(label="Анализ по конкурсам", callback = ex_viz_konk)
            dpg.add_menu_item(label="Анализ по субъектам фед.", callback = ex_viz_fed)
        with dpg.menu(label = "Финансирование"):
            dpg.add_menu_item(label = "Финансирование проектов", callback = finance_window)
        for i in window_specs:
            with dpg.window(tag = i[0], label  = i[0], width = 1200, height = 950, show = False, no_resize = True):
                dpg.set_item_pos(i[0],(400,50))
                with dpg.table(height = 800, label = i[0].lower(), header_row=True,
                            borders_innerH=True, borders_outerH=True, borders_innerV=True,
                            borders_outerV=True, resizable = False, scrollX=True, scrollY=True, no_keep_columns_visible=True, 
                            policy = dpg.mvTable_SizingFixedFit, sortable = True, callback=sort_callback, sort_multi=True):
                    table_name_id.append([dpg.last_item(), i[0].lower()])
                    for _label in eval( i[0] + '_label_list_ru'):
                        dpg.add_table_column(label=_label)
                    table_creation(table_list_name_size[i[1]])
                if i[0] == 'GR_PROJ':
                    with dpg.child_window(height = 100, autosize_x = True, autosize_y = False):
                            dpg.add_button(label = 'Добавить', pos = (100, 20), width = 200, height = 60, callback = add_row_btn)
                            dpg.add_button(label = 'Удалить', pos = (350,20), width = 200, height = 60, callback = delete_btn)
                            dpg.add_button(label = 'Редактировать', pos = (600,20), width = 200, height = 60, callback = edit_btn)
                            dpg.add_button(label = 'Фильтр', pos = (850,20), width = 200, height = 60, callback = filter_btn)         
    filter_condition_window()

dpg.setup_dearpygui()
dpg.show_viewport()
#dpg.maximize_viewport()
dpg.start_dearpygui()
dpg.destroy_context()